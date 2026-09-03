#!/usr/bin/env python3
"""Read the transport workbook into JSON rows for import_transport_workbook.php.

Replaces read_transport_workbook.ps1, which needed the Microsoft ACE.OLEDB
provider and skipped the "2025-2026" sheet.

One worksheet row becomes one load. Rows are never merged: booking_reference is
not unique in this business (a booking can be split across containers or part
shipments), and the RAIL sheet independently reuses the ROAD sheet's numbers.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ModuleNotFoundError:
    sys.exit("Missing dependency. Run: py -m pip install openpyxl")

# Worksheet title (normalized) -> sheet key used by the importer.
SHEETS = {
    "2025 2026": "master",
    "ROAD": "road",
    "AIR": "air",
    "FCL": "sea_fcl",
    "LCL": "sea_lcl",
    "RAIL": "rail",
    "CUSTOMS": "customs",
}

FIELDS = {
    "date": "DATE DATUM",
    "status": "SHIPMENT STATUS",
    "booking": "BOOKING REFERENCE",
    "insurance": "INSURANCE",
    "department": "DEPARTMENT",
    "freight_mode": "FREIGHT MODE",
    "consignee": "CONSIGNEE",
    "subdepartment": "SUBDEPARTMENT",
    "kgs": "KGS",
    "quantity": "QTY G W MEAS",
    "cbm": "CBM",
    "teu": "TEU",
    "container_types": "CONTAINER TYPES",
    "container": "CONTAINER",
    "departure": "DEPARTURE PORT STATION",
    "arrival": "ARRIVAL PORT STATION",
    "etd": "ETD DATE",
    "eta": "ETA DATE",
    "atd": "ATD DATE",
    "shipper": "SHIPPER NAME",
    "mediator": "MEDIATOR",
    "incoterms": "INCOTERMS",
    "price": "PRICE INSURANCE",
    "profit_loss": "GP PROFIT LOSS",
}

REQUIRED = ("booking", "consignee", "date")

# Consecutive blank rows that end a sheet scan.
BLANK_RUN_LIMIT = 200

# Canonical coordinates for the locations and transport codes used by the
# workbook. Distances are stored in kilometres and are deliberately left null
# when the source does not identify both endpoints (for example, "EXW").
LOCATION_COORDINATES = {
    "ATHENS": (37.9838, 23.7275),
    "BANJA LUKA": (44.7722, 17.1910),
    "BEOGRAD": (44.7866, 20.4489),
    "BG": (44.7866, 20.4489),
    "BIJELJINA": (44.7569, 19.2164),
    "BRCKO": (44.8728, 18.8083),
    "CHINA": (35.8617, 104.1954),
    "CNNGB": (29.8683, 121.5440),
    "CNTAO": (36.0671, 120.3826),
    "DE": (51.1657, 10.4515),
    "DUBRVONIK": (42.6507, 18.0944),
    "FOZHOU": (26.0745, 119.2965),
    "FUZU": (26.0745, 119.2965),
    "GRADACAC": (44.8785, 18.4276),
    "GUANGZHOU": (23.1291, 113.2644),
    "HRRJK": (45.3271, 14.4422),
    "ISTANBUL": (41.0082, 28.9784),
    "KLJUC": (44.5325, 16.7768),
    "KRUSEVAC": (43.5800, 21.3339),
    "LUKAVAC": (44.5425, 18.5262),
    "MODLNICZKA": (50.1290, 19.8640),
    "MUNDRA": (22.8395, 69.7213),
    "NANSHA": (22.7945, 113.5440),
    "NAVA SHEVA": (18.9497, 72.9510),
    "NHAVA SHEVA": (18.9497, 72.9510),
    "NEW YORK": (40.7128, -74.0060),
    "NINGBO": (29.8683, 121.5440),
    "NOVGRAD": (45.0464, 16.3778),
    "NOVI SAD": (45.2671, 19.8335),
    "NY": (40.7128, -74.0060),
    "PEK": (40.0799, 116.6031),
    "PETROVO": (44.6280, 18.3590),
    "PLOCE": (43.0560, 17.4310),
    "PODGORICA": (42.4304, 19.2594),
    "QINGDAO": (36.0671, 120.3826),
    "RIJEKA": (45.3271, 14.4422),
    "SANISKI MOST": (44.7667, 16.6670),
    "SANSKI MOST": (44.7667, 16.6670),
    "SARAJEVO": (43.8563, 18.4131),
    "SHANDONG": (36.3427, 118.1498),
    "SHANGAHI": (31.2304, 121.4737),
    "SHANGHAI": (31.2304, 121.4737),
    "SHENZHEN": (22.5431, 114.0579),
    "SHUNDE": (22.8050, 113.2930),
    "SJJ": (43.8246, 18.3315),
    "TIANJIN": (39.0842, 117.2009),
    "TUZLA": (44.5384, 18.6671),
    "TZ": (44.5384, 18.6671),
    "USA": (39.8283, -98.5795),
    "YANTIAN": (22.5560, 114.2370),
    "YANTINA": (22.5560, 114.2370),
}


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value if value is not None else ""))
    text = text.encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", re.sub(r"[^A-Za-z0-9]+", " ", text)).strip().upper()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y.")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def day_key(value: Any) -> str:
    """Normalize a date cell to YYYY-MM-DD; the sheets mix two formats."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    match = re.match(r"^(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})", text)
    if match:
        return f"{match.group(3)}-{int(match.group(2)):02d}-{int(match.group(1)):02d}"
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    return f"{match.group(1)}-{match.group(2)}-{match.group(3)}" if match else text


def coordinates(value: Any) -> tuple[float, float] | None:
    key = norm(value)
    # A slash sometimes identifies two possible endpoints, so it is not safe
    # to silently pick one of them.
    if not key or "/" in str(value):
        return None
    return LOCATION_COORDINATES.get(key)


def distance_km(departure: Any, arrival: Any) -> float | None:
    start = coordinates(departure)
    end = coordinates(arrival)
    if start is None or end is None:
        return None

    lat1, lon1 = map(math.radians, start)
    lat2, lon2 = map(math.radians, end)
    delta_lat = lat2 - lat1
    delta_lon = lon2 - lon1
    haversine = math.sin(delta_lat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lon / 2) ** 2

    return round(6371.0088 * 2 * math.asin(math.sqrt(haversine)), 2)


def header_row(worksheet) -> int | None:
    for row in range(1, min(7, worksheet.max_row or 0) + 1):
        for col in range(1, (worksheet.max_column or 0) + 1):
            if norm(worksheet.cell(row, col).value) == "BOOKING REFERENCE":
                return row
    return None


def read_sheet(worksheet, key: str) -> list[dict[str, Any]]:
    head = header_row(worksheet)
    if head is None:
        raise SystemExit(f"Booking reference header not found in sheet {worksheet.title!r}.")

    columns = {
        norm(worksheet.cell(head, col).value): col
        for col in range(1, worksheet.max_column + 1)
        if worksheet.cell(head, col).value not in (None, "")
    }
    missing = [FIELDS[f] for f in REQUIRED if FIELDS[f] not in columns]
    if missing:
        raise SystemExit(f"Sheet {worksheet.title!r} is missing required columns: {missing}")

    # Sheets carry formatting far below the data (ROAD reports ~1.05M rows), so
    # stop once the key columns have been blank for a long stretch.
    key_columns = [columns[FIELDS[f]] for f in REQUIRED]
    rows: list[dict[str, Any]] = []
    blank_run = 0
    for index in range(head + 1, worksheet.max_row + 1):
        if all(worksheet.cell(index, col).value in (None, "") for col in key_columns):
            blank_run += 1
            if blank_run >= BLANK_RUN_LIMIT:
                break
            continue
        blank_run = 0
        raw = {
            field: worksheet.cell(index, columns[header]).value if header in columns else None
            for field, header in FIELDS.items()
        }
        row = {field: cell_text(value) for field, value in raw.items()}
        if not all(row[f] for f in REQUIRED):
            continue
        row["sheet"] = key
        row["source_row"] = index
        row["date_key"] = day_key(raw["date"])
        row["distance_km"] = distance_km(raw["departure"], raw["arrival"])
        rows.append(row)
    return rows


def main() -> None:
    default = Path(__file__).resolve().parents[1] / "docs/Transport nalog - statusi 2026 BLS BH.xlsx"
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", type=Path, default=default)
    parser.add_argument("--year", type=int, default=2026, help="Keep only this year; 0 keeps every row.")
    parser.add_argument("--skip-sheet", action="append", default=[], metavar="KEY",
                        help="Sheet key to leave out, e.g. --skip-sheet master.")
    parser.add_argument("--report", action="store_true", help="Write a summary to stderr.")
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.path, data_only=True)
    rows: list[dict[str, Any]] = []
    found: set[str] = set()

    for worksheet in workbook.worksheets:
        key = SHEETS.get(norm(worksheet.title))
        if key is None:
            print(f"warning: ignoring unknown sheet {worksheet.title!r}", file=sys.stderr)
            continue
        found.add(key)
        if key in args.skip_sheet:
            continue
        rows.extend(read_sheet(worksheet, key))

    for expected in SHEETS.values():
        if expected not in found:
            print(f"warning: expected sheet {expected!r} not found in workbook", file=sys.stderr)

    dropped = 0
    if args.year:
        keep = [r for r in rows if r["date_key"].startswith(str(args.year))]
        dropped = len(rows) - len(keep)
        rows = keep

    rows.sort(key=lambda r: (r["date_key"], r["sheet"], r["source_row"]))

    if args.report:
        counts: dict[str, int] = {}
        for row in rows:
            counts[row["sheet"]] = counts.get(row["sheet"], 0) + 1
        print(f"rows emitted        : {len(rows)}", file=sys.stderr)
        print(f"rows dropped by year: {dropped}", file=sys.stderr)
        for sheet in sorted(counts):
            print(f"   {sheet:10} {counts[sheet]}", file=sys.stderr)

    json.dump(rows, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
